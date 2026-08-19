import pytest
import asyncio
from openpyxl import Workbook, load_workbook

from app import main


def test_office_path_blocks_escape(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())
    with pytest.raises(ValueError):
        main.safe_path("../../../etc/shadow")


def test_output_defaults_to_outputs(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())
    path = main.output_path("报告", ".docx")
    assert path == tmp_path / "outputs" / "报告.docx"


def test_template_escape_blocked(tmp_path, monkeypatch):
    templates = tmp_path / "templates"
    (templates / "word").mkdir(parents=True)
    outside = tmp_path / "secret.docx"
    outside.write_bytes(b"x")
    monkeypatch.setattr(main, "TEMPLATES", templates.resolve())
    with pytest.raises(ValueError):
        main.safe_template("word", "../../secret.docx")


def test_create_xlsx_reopens_file(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())

    async def fake_validate(path, kind):
        return {"validation": {kind: "PASS", "libreoffice_pdf": "PASS", "pdf_pages": 1}}

    monkeypatch.setattr(main, "validate_office", fake_validate)
    result = asyncio.run(main.create_xlsx({"path": "测试.xlsx", "sheets": {"汇总": [["项目", "数量"], ["A", 10]]}}))
    assert result["validation"]["openpyxl_reload"] == "PASS"
    book = load_workbook(tmp_path / "outputs" / "测试.xlsx")
    assert book["汇总"]["B2"].value == 10


def test_edit_and_analyze_xlsx(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())
    outputs = tmp_path / "outputs"
    outputs.mkdir()
    book = Workbook()
    book.active.title = "数据"
    book.active.append(["项目", "数量"])
    book.active.append(["A", 10])
    book.save(outputs / "source.xlsx")

    async def fake_validate(path, kind):
        return {"validation": {kind: "PASS", "libreoffice_pdf": "PASS", "pdf_pages": 1}}

    monkeypatch.setattr(main, "validate_office", fake_validate)
    edited = asyncio.run(main.edit_xlsx({"source": "outputs/source.xlsx", "output": "outputs/edited.xlsx", "set_cells": [{"sheet": "数据", "cell": "B2", "value": 20}]}))
    assert edited["ok"] is True
    analysis = asyncio.run(main.analyze_xlsx({"path": "outputs/edited.xlsx"}))
    assert analysis["analysis"]["数据"]["numeric"]["数量"]["sum"] == 20


def test_pdf_to_images(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())
    outputs = tmp_path / "outputs"
    outputs.mkdir()
    document = main.fitz.open()
    document.new_page()
    document.save(outputs / "one.pdf")
    document.close()
    result = asyncio.run(main.pdf_to_images({"path": "outputs/one.pdf", "dpi": 72}))
    assert result["page_count"] == 1
    assert (outputs / "one-pages" / "page-0001.png").exists()


def test_merge_xlsx_copies_styles_safely(tmp_path, monkeypatch):
    monkeypatch.setattr(main, "ROOT", tmp_path.resolve())
    outputs = tmp_path / "outputs"
    outputs.mkdir()
    for name, value in (("a.xlsx", 10), ("b.xlsx", 20)):
        book = Workbook()
        book.active.title = "数据"
        book.active.append(["项目", "数量"])
        book.active.append([name, value])
        book.active["A1"].font = main.Font(bold=True, color="FFFFFF")
        book.save(outputs / name)

    async def fake_validate(path, kind):
        return {"validation": {kind: "PASS", "libreoffice_pdf": "PASS", "pdf_pages": 1}}

    monkeypatch.setattr(main, "validate_office", fake_validate)
    result = asyncio.run(main.merge_xlsx({"path": "outputs/merged.xlsx", "sources": ["outputs/a.xlsx", "outputs/b.xlsx"]}))
    assert result["ok"] is True
    merged = load_workbook(outputs / "merged.xlsx")
    assert len(merged.sheetnames) == 2
    assert merged[merged.sheetnames[0]]["A1"].font.bold is True
